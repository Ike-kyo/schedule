# -*- coding: utf-8 -*-
import os
import tempfile
import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
import openpyxl
import win32com.client

DEFAULT_IMPORT = r"\\PC011\Users\yasumoku\Desktop\タカラ関係\工程表"
DEFAULT_OUTPUT = r"\\PC009\share01\日程表"

def convert_xls_to_xlsx(xls_path):
    excel = win32com.client.DispatchEx("Excel.Application")
    wb = excel.Workbooks.Open(xls_path)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    os.remove(tmp_path)
    wb.SaveAs(tmp_path, FileFormat=51)
    wb.Close(False)
    excel.Quit()
    return tmp_path

def read_excel_sheet(file_path, sheet_name):
    ext = os.path.splitext(file_path)[1].lower()
    tmp_xlsx = None
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        return wb, ws, tmp_xlsx
    elif ext == ".xls":
        tmp_xlsx = convert_xls_to_xlsx(file_path)
        wb = openpyxl.load_workbook(tmp_xlsx, data_only=True)
        ws = wb[sheet_name]
        return wb, ws, tmp_xlsx
    else:
        raise ValueError("対応していないファイル形式です")

def parse_excel_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y/%m/%d")
        except ValueError:
            return None
    return None

def normalize_key(val):
    if val is None:
        return ""
    return str(val).strip().replace("－", "-").replace("ー", "-").replace("−", "-")

def safe_int(v):
    """Excelセル値を安全に整数化する（空・文字列なら0）。"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except:
            return 0
    s = str(v).strip()
    if s == "":
        return 0
    s = s.replace(",", "")
    try:
        return int(float(s))
    except:
        return 0

def create_schedule(year, month, day, filter_type, import_path=DEFAULT_IMPORT, output_path=DEFAULT_OUTPUT, gui_select_file_func=None):
    """
    - filter_type: "all" または "dollar"（新図面のみ）
    - import_path: フォルダパス（末尾は自動調整）
    - gui_select_file_func: 複数候補のときにGUI側で選ばせるコールバック
    """
    import_path = os.path.normpath(import_path.rstrip("\\"))
    output_path = os.path.normpath(output_path.rstrip("\\"))

    target_date = datetime(year, month, day)

    base_name1 = f"{month}-{day}"
    base_name2 = f"{month:02d}-{day:02d}"
    candidates = []

    for f in os.listdir(import_path):
        if f.endswith((".xls", ".xlsx")) and (base_name1 in f or base_name2 in f):
            full_path = os.path.join(import_path, f)
            full_path = os.path.normpath(full_path)
            update_time = datetime.fromtimestamp(os.path.getmtime(full_path))
            candidates.append((f, update_time, full_path))
    if not candidates:
        raise FileNotFoundError(f"参照フォルダが存在しないかアクセスできません:{os.path.join(import_path, base_name1+'*.xls')}")

    if len(candidates) > 1 and gui_select_file_func is not None:
        choice = gui_select_file_func([(c[0], c[1]) for c in candidates])
        if not choice:
            raise Exception("行程表の選択がキャンセルされました")
        if os.path.isabs(choice) and os.path.exists(choice):
            schedule_file = choice
        else:
            schedule_file = os.path.join(import_path, os.path.basename(choice))
    else:
        schedule_file = candidates[0][2]

    wb_schedule, ws_schedule, tmp_schedule = read_excel_sheet(schedule_file, 'Sheet1')

    REQUEST_FILE = os.path.normpath(r"\\PC009\share01\依頼現場名 R1.xls")
    wb_req, ws_req, tmp_request = read_excel_sheet(REQUEST_FILE, "Sheet3")

    SCHEDULE_TEMPLATE = os.path.normpath(r"\\PC009\share01\日程表\生産日程表★.xlsx")
    wb_out = openpyxl.load_workbook(SCHEDULE_TEMPLATE)
    ws_out = wb_out.active

    today = datetime.today()
    start_date = today - relativedelta(months=5)
    end_date = today + relativedelta(months=2)

    written_rows = set()
    out_row = 2

    blocks = [(3, 62, "C"), (3, 62, "I")]
    for start_row, end_row, col_c in blocks:
        for r in range(start_row, end_row + 1):
            key_val = ws_schedule[f"{col_c}{r}"].value
            if key_val is None:
                continue
            key_val_str = normalize_key(key_val)
            for req_r in range(2, ws_req.max_row + 1):
                req_key = ws_req[f"C{req_r}"].value
                if req_key is None:
                    continue
                req_key_str = normalize_key(req_key)
                if key_val_str != req_key_str:
                    continue
                req_g = parse_excel_date(ws_req[f"G{req_r}"].value)
                if req_g is None or req_g < start_date or req_g > end_date:
                    continue
                val_d = str(ws_req[f"D{req_r}"].value or "").strip()
                val_c = str(ws_req[f"C{req_r}"].value or "").strip()

                if filter_type == "dollar":
                    if not (val_d.startswith("$") or val_d.startswith("＄")):
                        continue

                row_key = tuple(ws_req[f"{col}{req_r}"].value for col in "ABCDEF")
                if row_key in written_rows:
                    continue
                written_rows.add(row_key)

                val_a = str(ws_req[f"C{req_r}"].value or "").strip()
                ws_out[f"A{out_row}"] = val_a
                ws_out[f"A{out_row}"].number_format = '@'
                ws_out[f"C{out_row}"] = ws_req[f"A{req_r}"].value
                ws_out[f"F{out_row}"] = ws_req[f"B{req_r}"].value
                ws_out[f"E{out_row}"] = ws_req[f"D{req_r}"].value
                ws_out[f"D{out_row}"] = ws_req[f"E{req_r}"].value
                ws_out[f"B{out_row}"] = target_date.strftime("%Y/%m/%d")
                ws_out[f"I{out_row}"] = ws_req[f"F{req_r}"].value
                out_row += 1

    for r_out in range(2, out_row):
        val_A = normalize_key(ws_out[f"A{r_out}"].value)
        val_D = ws_out[f"D{r_out}"].value
        val_F = ws_out[f"F{r_out}"].value

        for r_schedule in range(3, 63):
            c = normalize_key(ws_schedule[f"C{r_schedule}"].value)
            d = ws_schedule[f"D{r_schedule}"].value
            e = ws_schedule[f"E{r_schedule}"].value
            f_val = ws_schedule[f"F{r_schedule}"].value
            if c == val_A and d == val_F and e == val_D:
                ws_out[f"I{r_out}"] = f_val
                break

        for r_schedule in range(3, 63):
            i = normalize_key(ws_schedule[f"I{r_schedule}"].value)
            j = ws_schedule[f"J{r_schedule}"].value
            k = ws_schedule[f"K{r_schedule}"].value
            l_val = ws_schedule[f"L{r_schedule}"].value
            if i == val_A and j == val_F and k == val_D:
                ws_out[f"I{r_out}"] = l_val
                break

    # --- 全件集計（メッセージ用） ---
    gifu_new_total = shiga_new_total = gifu_old_total = shiga_old_total = 0

    for r in range(3, 63):
        c = ws_schedule[f"C{r}"].value
        d = ws_schedule[f"D{r}"].value
        f_val = safe_int(ws_schedule[f"F{r}"].value)
        if c is None:
            continue
        first = str(c).strip()[:1]
        if re.match(r'[A-Za-z]', first):
            if isinstance(d, str) and (d.startswith("$") or d.startswith("＄")):
                gifu_new_total += f_val
            else:
                gifu_old_total += f_val
        else:
            if isinstance(d, str) and (d.startswith("$") or d.startswith("＄")):
                shiga_new_total += f_val
            else:
                shiga_old_total += f_val

    for r in range(3, 63):
        i = ws_schedule[f"I{r}"].value
        j = ws_schedule[f"J{r}"].value
        l_val = safe_int(ws_schedule[f"L{r}"].value)
        if i is None:
            continue
        first = str(i).strip()[:1]
        if re.match(r'[A-Za-z]', first):
            if isinstance(j, str) and (j.startswith("$") or j.startswith("＄")):
                gifu_new_total += l_val
            else:
                gifu_old_total += l_val
        else:
            if isinstance(j, str) and (j.startswith("$") or j.startswith("＄")):
                shiga_new_total += l_val
            else:
                shiga_old_total += l_val

    # --- 出力用集計（filter_typeに応じて） ---
    # 新図面のみの場合でも出力メッセージは全件集計を使う
    gifu_new = gifu_new_total
    shiga_new = shiga_new_total
    gifu_old = gifu_old_total
    shiga_old = shiga_old_total

    save_file = os.path.join(output_path, f"{year:04d}{month:02d}{day:02d}_生産日程表★.xlsx")
    save_file = os.path.normpath(save_file)
    wb_out.save(save_file)

    wb_out.close()
    wb_schedule.close()
    wb_req.close()
    if tmp_schedule and os.path.exists(tmp_schedule):
        os.remove(tmp_schedule)
    if tmp_request and os.path.exists(tmp_request):
        os.remove(tmp_request)

    return save_file, int(gifu_new), int(shiga_new), int(gifu_old), int(shiga_old)
